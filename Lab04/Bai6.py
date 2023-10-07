from openpyxl import *
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox
import re

wb = load_workbook('C:\\Users\\NHA\\Desktop\\form_sv.xlsx')

sheet = wb.active

def excel():
    sheet.column_dimensions["A"].width=30
    sheet.column_dimensions["B"].width=50
    sheet.column_dimensions["C"].width=10
    sheet.column_dimensions["D"].width=20
    sheet.column_dimensions["E"].width=10
    sheet.column_dimensions["F"].width=10
    sheet.column_dimensions["G"].width=20
    sheet.column_dimensions["H"].width=100


    sheet.cell(row=1, column=1).value="MSSV"
    sheet.cell(row=1, column=2).value="Họ tên"
    sheet.cell(row=1, column=3).value="Ngày sinh"
    sheet.cell(row=1, column=4).value="Email"
    sheet.cell(row=1, column=5).value="Số ĐT"
    sheet.cell(row=1, column=6).value="Học kỳ"
    sheet.cell(row=1, column=7).value="Năm học"
    sheet.cell(row=1, column=8).value="Môn học"

def clear():
    mssv_field.delete(0, END)
    hocten_field.delete(0, END)
    ngaysinh_field.delete(0, END)
    email_field.delete(0, END)
    sdt_field.delete(0, END)
    hocky_field.delete(0, END)
    namhoc_field.delete(0, END)

def insert():
    if (mssv_field.get() == "" and
       hocten_field.get() == "" and
       ngaysinh_field.get() == "" and
       email_field.get() == "" and
       sdt_field.get() == "" and
       hocky_field.get() == "" and
       namhoc_field.get() == ""):
        print("Empty input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
        sheet.cell(row=current_row + 1, column=2).value = hocten_field.get()
        sheet.cell(row=current_row + 1, column=3).value = ngaysinh_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = sdt_field.get()
        sheet.cell(row=current_row + 1, column=6).value = hocky_field.get()
        sheet.cell(row=current_row + 1, column=7).value = namhoc_field.get()
        sheet.cell(row=current_row + 1, column=8).value = chonMon1.get()


        wb.save('C:\\Users\\NHA\\Desktop\\form_sv.xlsx')

        mssv_field.focus_set()

        clear()

special_ch = ['~', '`', '!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '-', '_', '+', '=', '{', '}', '[', ']', '|', '\\', '/', ':', ';', '"', "'", '<', '>', ',', '.', '?']

def KtraMS():
    ms = mssv_field.get()
    maSo = ""
    if len(ms) == 0:
        maSo = 'Ma so khong duoc de trong!'
    else:
        try:
            if any(tu in special_ch for tu in ms):
                maSo = 'Khong duoc phep nhap ky tu'
            elif any(tu.isupper() or tu.islower() for tu in ms):
                maSo = 'Khong duoc phep nhap chu'
            elif len(ms) < 7:
                maSo = 'Ma so phai co do dai bang 7!'
        except Exception as ep:
            messagebox.showerror('error', ep)
    messagebox.showinfo('Thong Bao', maSo)  

def KtraMail():
    mail = email_field.get()
    regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    if not re.match(regex, mail):
        email_field.delete(0, END)
        messagebox.showerror(0, "Loi dinh dang Email")

def KtraSDT():
    soDT = sdt_field.get()
    dt = ""
    if len(soDT) == 0:
        dt = 'So dien thoai khong duoc de trong!'
    else:
        try:
            if len(soDT) < 10:
                dt = 'So dien thoai phai co do dai bang 10!'
        except Exception as ep:
            messagebox.showerror('error', ep)
    messagebox.showinfo('Thong Bao', dt)

def KtraHocKy():
    hocKy = hocky_field.get()
    if hocKy not in ['1','2','3']:
        hocky_field.delete(0, END)
        messagebox.showerror(0,"Chi duoc nhap vao hoc ky (1, 2, 3)")

def KtraNgay():
    ngay = ngaysinh_field.get()
    dk = r'^\d{2}/\d{2}/\d{4}$'
    if not re.match(dk, ngay):
        ngaysinh_field.delete(0, END)
        messagebox.showerror(0, "Dinh dang ngay sai (dd/mm/yyyy)")

if __name__ == "__main__":
    #Create Layout
    root = Tk()
    root.configure(background="light green")
    root.geometry("580x300")
    root.title("Đăng ký học phần")


    #create Label
    heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green", fg="red", font=('sanif', 12))
    mssv = Label(root, text="Mã số sinh viên", bg="light green")
    hoten = Label(root, text="Họ tên", bg="light green")
    ngaysinh = Label(root, text="Ngày sinh", bg="light green")
    email = Label(root, text="Email", bg="light green")
    sdt = Label(root, text="Số điện thoại", bg="light green")
    hocky = Label(root, text="Học kỳ", bg="light green")
    namhoc = Label(root, text="Năm học", bg="light green")
    chonmonhoc = Label(root, text="Chọn môn học", bg="light green")

    heading.place(anchor='center', relx=0.5, rely=0.1)
    mssv.place(relx=0.01, rely=0.15)
    hoten.place(relx=0.01, rely=0.22)
    ngaysinh.place(relx=0.01, rely=0.29)
    email.place(relx=0.01, rely=0.36)
    sdt.place(relx=0.01, rely=0.43)
    hocky.place(relx=0.01, rely=0.50)
    chonmonhoc.place(relx=0.01, rely=0.64)

    #create entry
    mssv_field = Entry(root)
    hocten_field = Entry(root)
    ngaysinh_field = Entry(root)
    email_field = Entry(root)
    sdt_field = Entry(root)
    hocky_field = Entry(root)
    namhoc_field = Entry(root)
    nh = StringVar()
    namhoc_field = Combobox(root, width=57, textvariable=nh)
    namhoc_field['values'] = ('2022-2023', '2023-2024', '2024-2025')
    namhoc_field.current(0)
    namhoc_field.place(relx=0.18, rely=0.57, width=300)

    chonMon1 = StringVar()
    chonMon2= StringVar()
    chonMon3 = StringVar()
    chonMon4 = StringVar()
    chonMon_check = Checkbutton(root,text="Lập trình Python",bg="light green",variable=chonMon1).place(relx=0.2, rely=0.7)
    chonMon_check = Checkbutton(root,text="Lập trình Java",bg="light green",variable=chonMon2).place(relx=0.2, rely=0.8)
    chonMon_check = Checkbutton(root,text="Công nghệ phần mềm",bg="light green", variable=chonMon3).place(relx=0.43, rely=0.7)
    chonMon_check = Checkbutton(root,text="Phát triển ứng dụng Web",bg="light green",variable=chonMon4).place(relx=0.43, rely=0.8)
    

    mssv_field.place(relx=0.18, rely=0.15, width=300)
    hocten_field.place(relx=0.18, rely=0.22, width=300)
    ngaysinh_field.place(relx=0.18, rely=0.29, width=300)
    email_field.place(relx=0.18, rely=0.36, width=300)
    sdt_field.place(relx=0.18, rely=0.43, width=300)
    hocky_field.place(relx=0.18, rely=0.50, width=300)
    namhoc_field.place(relx=0.18, rely=0.57, width=300)

    excel()

    btnDky = Button(root, text="Đăng ký", bg="green", command=lambda:[KtraMS(), KtraMail(), KtraHocKy(), KtraNgay() ,KtraSDT(), insert()])
    btnDky.place(relx=0.52, rely=0.9)

    btnExit = Button(root, text="Thoát", bg="green", command=quit)
    btnExit.place(relx=0.62, rely=0.9)

    root.mainloop()