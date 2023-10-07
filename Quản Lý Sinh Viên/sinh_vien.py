from openpyxl import *
from tkinter import *
from tkinter import messagebox, Radiobutton
from tkinter.ttk import Combobox
import re
from tkinter import ttk

wb = load_workbook('C:\\Users\\NHA\\Desktop\\form_sv.xlsx')

sheet = wb.active
selected_row = None

def excel():
    sheet.column_dimensions["A"].width=30
    sheet.column_dimensions["B"].width=50
    sheet.column_dimensions["C"].width=10
    sheet.column_dimensions["D"].width=20
    sheet.column_dimensions["E"].width=10
    sheet.column_dimensions["F"].width=10
    sheet.column_dimensions["G"].width=20
    sheet.column_dimensions["H"].width=100


    sheet.cell(row=1, column=1).value="MSSV"
    sheet.cell(row=1, column=2).value="Họ tên"
    sheet.cell(row=1, column=3).value="Ngày sinh"
    sheet.cell(row=1, column=4).value="Email"
    sheet.cell(row=1, column=5).value="Số ĐT"
    sheet.cell(row=1, column=6).value="Học kỳ"
    sheet.cell(row=1, column=7).value="Năm học"
    sheet.cell(row=1, column=8).value="Môn học"

def clear():
    mssv_field.delete(0, END)
    hocten_field.delete(0, END)
    ngaysinh_field.delete(0, END)
    email_field.delete(0, END)
    sdt_field.delete(0, END)
    hocky_field.delete(0, END)
    namhoc_field.delete(0, END)

def insert():
    global selected_row

    if (mssv_field.get() == "" 
            and hocten_field.get() == ""
            and ngaysinh_field.get() == ""
            and email_field.get() == ""
            and sdt_field.get() == ""
            and hocky_field.get() == ""
            and namhoc_field.get() == ""):
            print("Empty input")
    elif selected_row:
            index = student_tree.index(selected_row)
            current_row = index + 2

            student_tree.item(selected_row, values=(
                mssv_field.get(),
                hocten_field.get(),
                ngaysinh_field.get(),
                email_field.get(),
                sdt_field.get(),
                hocky_field.get(),
                namhoc_field.get(),
                chonMon1.get()
            ))

            sheet.cell(row=current_row, column=1).value = mssv_field.get()
            sheet.cell(row=current_row, column=2).value = hocten_field.get()
            sheet.cell(row=current_row, column=3).value = ngaysinh_field.get()
            sheet.cell(row=current_row, column=4).value = email_field.get()
            sheet.cell(row=current_row, column=5).value = sdt_field.get()
            sheet.cell(row=current_row, column=6).value = hocky_field.get()
            sheet.cell(row=current_row, column=7).value = namhoc_field.get()
            sheet.cell(row=current_row, column=8).value = chonMon1.get()

            wb.save('C:\\Users\\NHA\\Desktop\\form_sv.xlsx')

            mssv_field.focus_set()

            update_treeview()

            clear()
    else:
            current_row = sheet.max_row +1
            current_column = sheet.max_column

            sheet.cell(row=current_row , column=1).value = mssv_field.get()
            sheet.cell(row=current_row , column=2).value = hocten_field.get()
            sheet.cell(row=current_row , column=3).value = ngaysinh_field.get()
            sheet.cell(row=current_row , column=4).value = email_field.get()
            sheet.cell(row=current_row , column=5).value = sdt_field.get()
            sheet.cell(row=current_row , column=6).value = hocky_field.get()
            sheet.cell(row=current_row , column=7).value = namhoc_field.get()
            sheet.cell(row=current_row , column=8).value = chonMon1.get()


            wb.save('C:\\Users\\NHA\\Desktop\\form_sv.xlsx')

            mssv_field.focus_set()

            update_treeview()

            clear()
    

special_char = ['~', '`', '!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '-', '_', '+', '=', '{', '}', '[', ']', '|', '\\', '/', ':', ';', '"', "'", '<', '>', ',', '.', '?']

def ktra_mssv():
    ms = mssv_field.get()
    maSo = ""
    if len(ms) == 0:
        maSo = 'Ma so khong duoc de trong!'
    else:
        try:
            if any(tu in special_char for tu in ms):
                maSo = 'Khong duoc phep nhap ky tu'
            elif any(tu.isupper() or tu.islower() for tu in ms):
                maSo = 'Khong duoc phep nhap chu'
            elif len(ms) < 7:
                maSo = 'Ma so phai co do dai bang 7!'
        except Exception as ep:
            messagebox.showerror('error', ep)
    messagebox.showinfo('Thong Bao', maSo)  

def ktra_email():
    mail = email_field.get()
    regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    if not re.match(regex, mail):
        email_field.delete(0, END)
        messagebox.showerror(0, "Loi dinh dang Email")

def ktra_sdt():
    soDT = sdt_field.get()
    dt = ""
    if len(soDT) == 0:
        dt = 'So dien thoai khong duoc de trong!'
    else:
        try:
            if len(soDT) < 10:
                dt = 'So dien thoai phai co do dai bang 10!'
        except Exception as ep:
            messagebox.showerror('error', ep)
    messagebox.showinfo('Thong Bao', dt)

def ktra_hocky():
    hocKy = hocky_field.get()
    if hocKy not in ['1','2','3']:
        hocky_field.delete(0, END)
        messagebox.showerror(0,"Chi duoc nhap vao hoc ky (1, 2, 3)")

def ktra_ngay():
    ngay = ngaysinh_field.get()
    dk = r'^\d{2}/\d{2}/\d{4}$'
    if not re.match(dk, ngay):
        ngaysinh_field.delete(0, END)
        messagebox.showerror(0, "Dinh dang ngay sai (dd/mm/yyyy)")

def update_treeview():
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        student_tree.insert("", "end", values=row)

def delete_student():
    global selected_row
    
    selected_item = student_tree.selection()
    
    if not selected_item:
        messagebox.showwarning("Thông báo", "Hãy chọn một sinh viên để xóa.")
        return
    
    confirmation = messagebox.askyesno("Xác nhận xóa", "Bạn có chắc chắn muốn xóa sinh viên này?")
    
    if confirmation:
        id_to_delete = student_tree.item(selected_item, "values")[0]
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == id_to_delete:
                sheet.delete_rows(row[0].row)
                break
        
        wb.save('C:\\Users\\NHA\\Desktop\\form_sv.xlsx')
        
        update_treeview()

        reset_treeview()

        clear()
        
        selected_row = None


def search_student():
    for row in student_tree.get_children():
        student_tree.delete(row)

    option = search_option.get()
    search_text = search_entry.get()

    if option == "ID":
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[0] == search_text:
                student_tree.insert("", "end", values=row)
    
    elif option == "Tên":
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if search_text.lower() in row[1].lower():
                student_tree.insert("", "end", values=row)

def reset_treeview():
    for row in student_tree.get_children():
        student_tree.delete(row)
    
    clear()
    update_treeview()

def on_treeview_click(event):
    global selected_row

    selected_row = student_tree.selection()[0]
    values = student_tree.item(selected_row, 'values')

    mssv_field.delete(0, END)
    mssv_field.insert(0, values[0])

    hocten_field.delete(0, END)
    hocten_field.insert(0, values[1])

    ngaysinh_field.delete(0, END)
    ngaysinh_field.insert(0, values[2])

    email_field.delete(0, END)
    email_field.insert(0, values[3])

    sdt_field.delete(0, END)
    sdt_field.insert(0, values[4])

    hocky_field.delete(0, END)
    hocky_field.insert(0, values[5])

    nh.set(values[6])
    chonMon1.set(values[7])


if __name__ == "__main__":
    root = Tk()
    root.configure(background="light sky blue")
    root.geometry("1080x650")
    root.title("Quản Lý Sinh Viên")


    heading = Label(root, text="QUẢN LÝ SINH VIÊN", bg="dodger blue", fg="white", font=('sanif', 18, 'bold'))
    mssv = Label(root, text="Mã số sinh viên", bg="light sky blue")
    hoten = Label(root, text="Họ tên", bg="light sky blue")
    ngaysinh = Label(root, text="Ngày sinh", bg="light sky blue")
    email = Label(root, text="Email", bg="light sky blue")
    sdt = Label(root, text="Số điện thoại", bg="light sky blue")
    hocky = Label(root, text="Học kỳ", bg="light sky blue")
    namhoc = Label(root, text="Năm học", bg="light sky blue")
    chonmonhoc = Label(root, text="Chọn môn học", bg="light sky blue")

    headtim = Label(text="Tìm kiếm", fg="white",  bg="light sky blue", font=("sanif", 14, "bold"))
    headtim.place(relx=0.5, rely=0.1)

    heading.place(anchor='center', relx=0.5, rely=0.04)
    mssv.place(relx=0.01, rely=0.1)
    hoten.place(relx=0.01, rely=0.15)
    ngaysinh.place(relx=0.01, rely=0.20)
    email.place(relx=0.01, rely=0.25)
    sdt.place(relx=0.01, rely=0.3)
    hocky.place(relx=0.01, rely=0.35)
    namhoc.place(relx=0.01, rely=0.4)
    chonmonhoc.place(relx=0.01, rely=0.45)

    mssv_field = Entry(root)
    hocten_field = Entry(root)
    ngaysinh_field = Entry(root)
    email_field = Entry(root)
    sdt_field = Entry(root)
    hocky_field = Entry(root)
    namhoc_field = Entry(root)
    nh = StringVar()
    namhoc_field = Combobox(root, width=57, textvariable=nh)
    namhoc_field['values'] = ('2022-2023', '2023-2024', '2024-2025')
    namhoc_field.current(0)

    chonMon1 = StringVar()
    chonMon2= StringVar()
    chonMon3 = StringVar()
    chonMon4 = StringVar()
    chonMon_check = Checkbutton(root,text="Lập trình Python",bg="light green",variable=chonMon1).place(relx=0.1, rely=0.45)
    chonMon_check = Checkbutton(root,text="Lập trình Java",bg="light green",variable=chonMon2).place(relx=0.1, rely=0.5)
    chonMon_check = Checkbutton(root,text="Công nghệ phần mềm",bg="light green", variable=chonMon3).place(relx=0.25, rely=0.45)
    chonMon_check = Checkbutton(root,text="Phát triển ứng dụng Web",bg="light green",variable=chonMon4).place(relx=0.25, rely=0.5)
    

    mssv_field.place(relx=0.1, rely=0.1, width=300)
    hocten_field.place(relx=0.1, rely=0.15, width=300)
    ngaysinh_field.place(relx=0.1, rely=0.2, width=300)
    email_field.place(relx=0.1, rely=0.25, width=300)
    sdt_field.place(relx=0.1, rely=0.3, width=300)
    hocky_field.place(relx=0.1, rely=0.35, width=300)
    namhoc_field.place(relx=0.1, rely=0.4, width=300)

    student_tree = ttk.Treeview(root, columns=("MSSV", "Họ tên", "Ngày sinh", "Email", "Số ĐT", "Học kỳ", "Năm học", "Môn học"))
    student_tree.place(relx=0.01, rely=0.55, relwidth=0.98, relheight=0.35)

    student_tree.heading("MSSV", text="MSSV")
    student_tree.heading("Họ tên", text="Họ tên")
    student_tree.heading("Ngày sinh", text="Ngày sinh")
    student_tree.heading("Email", text="Email")
    student_tree.heading("Số ĐT", text="Số ĐT")
    student_tree.heading("Học kỳ", text="Học kỳ")
    student_tree.heading("Năm học", text="Năm học")
    student_tree.heading("Môn học", text="Môn học")

    student_tree.column("MSSV", width=70)
    student_tree.column("Họ tên", width=150)
    student_tree.column("Ngày sinh", width=80)
    student_tree.column("Email", width=150)
    student_tree.column("Số ĐT", width=80)
    student_tree.column("Học kỳ", width=50)
    student_tree.column("Năm học", width=80)
    student_tree.column("Môn học", width=200)
    student_tree.bind("<ButtonRelease-1>", on_treeview_click)

    search_option = StringVar()
    search_option.set("MSSV")

    search_by_id_radio = Radiobutton(root, text="Tìm theo MSSV", variable=search_option, value="ID")
    search_by_name_radio = Radiobutton(root, text="Tìm theo tên", variable=search_option, value="Tên")
    search_by_id_radio.place(relx=0.5, rely=0.15)
    search_by_name_radio.place(relx=0.6, rely=0.15)

    search_entry = Entry(root)
    search_entry.place(relx=0.5, rely=0.2, width=200, height=20)

    search_button = Button(root, text="Tìm kiếm", bg="blue", command=search_student)
    search_button.place(relx=0.61, rely=0.24, width=80)

    reset_button = Button(root, text="Reset", bg="red", command=reset_treeview)
    reset_button.place(relx=0.5, rely=0.24, width=80)

    excel()

    delete_button = Button(root, text="Xóa", bg="red", command=delete_student)
    delete_button.place(relx=0.7, rely=0.92, width=100)
    
    btnDky = Button(root, text="Thêm", bg="green", command=lambda:[ktra_mssv(), ktra_email(), ktra_hocky(), ktra_ngay() ,ktra_sdt(), insert()])
    btnDky.place(relx=0.8, rely=0.92, width=100)

    btnExit = Button(root, text="Thoát", bg="green", command=quit)
    btnExit.place(relx=0.9, rely=0.92, width=100)

    root.mainloop()
